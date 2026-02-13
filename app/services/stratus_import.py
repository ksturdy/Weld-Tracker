import re
from openpyxl import load_workbook
from app import db
from app.models import StratusMap, WeldEntry


def import_stratus_map(xlsx_path):
    """Import a Stratus map Excel file that maps weld labels to package numbers.

    Handles two known sheet formats:
    - DH1100: Col 0=Weld Label, Col 3=Package Number
    - DH1300: Col 0=Weld Label, Col 1=Package, Col 2=Size, Col 3=Type, Col 4=Description

    Returns a summary dict with counts.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    summary = {'sheets_processed': 0, 'entries_imported': 0, 'entries_updated': 0}

    # Load all existing labels in one query to avoid per-row lookups
    existing_map = {s.weld_label: s for s in StratusMap.query.all()}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            continue

        # Detect the column layout by finding the header row
        col_map = _detect_columns(rows, sheet_name)
        if not col_map:
            continue

        summary['sheets_processed'] += 1
        header_row = col_map.pop('_header_row', 1)

        for row in rows[header_row + 1:]:
            if not row or row[col_map['label']] is None:
                continue

            label = str(row[col_map['label']]).strip()
            if not label or not re.match(r'^TW[A-Z]-\d+', label):
                continue

            package = ''
            if col_map.get('package') is not None and row[col_map['package']]:
                package = str(row[col_map['package']]).strip()

            size = ''
            if col_map.get('size') is not None and len(row) > col_map['size'] and row[col_map['size']]:
                size = str(row[col_map['size']]).strip()

            weld_type = ''
            if col_map.get('type') is not None and len(row) > col_map['type'] and row[col_map['type']]:
                weld_type = str(row[col_map['type']]).strip()

            description = ''
            if col_map.get('desc') is not None and len(row) > col_map['desc'] and row[col_map['desc']]:
                description = str(row[col_map['desc']]).strip()

            # Upsert using in-memory lookup (no per-row DB query)
            existing = existing_map.get(label)
            if existing:
                existing.package_number = package
                existing.size = size
                existing.weld_type = weld_type
                existing.description = description
                existing.sheet_name = sheet_name
                summary['entries_updated'] += 1
            else:
                entry = StratusMap(
                    weld_label=label,
                    package_number=package,
                    size=size,
                    weld_type=weld_type,
                    description=description,
                    sheet_name=sheet_name,
                )
                db.session.add(entry)
                existing_map[label] = entry
                summary['entries_imported'] += 1

    db.session.commit()
    wb.close()
    return summary


def _detect_columns(rows, sheet_name):
    """Detect column mapping from header rows."""
    # Try rows 0, 1, 2 as potential header rows
    for row_idx in range(min(3, len(rows))):
        row = rows[row_idx]
        if not row:
            continue

        headers = [str(h).strip().lower() if h else '' for h in row]

        col_map = {'_header_row': row_idx}

        for i, h in enumerate(headers):
            if 'weld label' in h or h == 'weld label per spool':
                col_map['label'] = i
            elif h in ('package number', 'package'):
                col_map['package'] = i
            elif h == 'size':
                col_map['size'] = i
            elif h == 'type':
                col_map['type'] = i
            elif h == 'description':
                col_map['desc'] = i

        if 'label' in col_map and 'package' in col_map:
            return col_map

    return None


def lookup_package(iso_number, weld_number):
    """Look up package number from the Stratus map.

    Args:
        iso_number: e.g. "TWS-189"
        weld_number: e.g. "Weld 1" or "1"

    Returns:
        Package number string or empty string if not found.
    """
    # Build the full weld label: TWS-189 + Weld 1 -> TWS-189-1
    m = re.search(r'(\d+)', str(weld_number)) if weld_number else None
    if m:
        weld_label = f'{iso_number}-{m.group(1)}'
    else:
        weld_label = iso_number

    entry = StratusMap.query.filter_by(weld_label=weld_label).first()
    return entry.package_number if entry else ''


def assign_packages_to_entries():
    """Bulk-assign package numbers to all weld entries that don't have one yet."""
    entries = WeldEntry.query.filter(
        (WeldEntry.package_number == None) | (WeldEntry.package_number == '')
    ).all()

    updated = 0
    for entry in entries:
        package = lookup_package(entry.iso_number, entry.weld_number)
        if package:
            entry.package_number = package
            updated += 1

    db.session.commit()
    return updated
