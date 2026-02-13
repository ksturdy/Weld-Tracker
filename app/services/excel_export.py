from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.models import InspectionReport, WeldEntry, UTDetail, UTIndication


def export_to_excel(filters=None):
    """Export weld data to an Excel workbook.

    Args:
        filters: dict with optional keys 'inspection_type', 'date_from', 'date_to',
                 'project', 'result'

    Returns:
        BytesIO buffer containing the Excel file.
    """
    wb = Workbook()

    _create_summary_sheet(wb, filters)
    _create_vt_sheet(wb, filters)
    _create_ut_sheet(wb, filters)

    # Remove default sheet if others exist
    if len(wb.sheetnames) > 1 and 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _build_query(filters):
    """Build base query with filters applied."""
    query = InspectionReport.query

    if filters:
        if filters.get('inspection_type'):
            query = query.filter(InspectionReport.inspection_type == filters['inspection_type'])
        if filters.get('date_from'):
            query = query.filter(InspectionReport.inspection_date >= filters['date_from'])
        if filters.get('date_to'):
            query = query.filter(InspectionReport.inspection_date <= filters['date_to'])
        if filters.get('project'):
            query = query.filter(InspectionReport.project_name.ilike(f"%{filters['project']}%"))
        if filters.get('result'):
            query = query.filter(InspectionReport.result_summary == filters['result'])

    return query.order_by(InspectionReport.inspection_date.desc())


def _style_header(ws, row=1):
    """Apply header styling to the first row."""
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border


def _create_summary_sheet(wb, filters):
    """Create summary sheet with all weld entries."""
    ws = wb.active
    ws.title = 'All Weld Entries'

    headers = [
        'ISO #', 'Weld #', 'Package', 'Inspection Type', 'Date', 'Result',
        'Job Number', 'Project', 'Client', 'Location', 'Inspector',
        'Specification', 'PDF File'
    ]
    ws.append(headers)
    _style_header(ws)

    accept_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    reject_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    reports = _build_query(filters).all()
    for report in reports:
        for entry in report.weld_entries:
            row = [
                entry.iso_number,
                entry.weld_number,
                entry.package_number or '',
                report.inspection_type,
                report.inspection_date.strftime('%m/%d/%Y') if report.inspection_date else '',
                entry.evaluation,
                report.job_number,
                report.project_name,
                report.client,
                report.location,
                report.inspector_name,
                report.specification,
                report.pdf_filename,
            ]
            ws.append(row)

            # Color-code result
            result_cell = ws.cell(row=ws.max_row, column=6)
            if entry.evaluation and 'accept' in entry.evaluation.lower():
                result_cell.fill = accept_fill
            elif entry.evaluation and 'reject' in entry.evaluation.lower():
                result_cell.fill = reject_fill

    # Auto-fit column widths
    for col in ws.columns:
        max_length = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)


def _create_vt_sheet(wb, filters):
    """Create VT-specific sheet."""
    vt_filters = dict(filters or {})
    vt_filters['inspection_type'] = 'VT'
    reports = _build_query(vt_filters).all()

    if not reports:
        return

    ws = wb.create_sheet('VT Reports')
    headers = [
        'ISO #', 'Weld #', 'Result', 'Date', 'Job Number',
        'Project', 'Client', 'Location', 'Equipment', 'Inspector', 'CWI #'
    ]
    ws.append(headers)
    _style_header(ws)

    for report in reports:
        for entry in report.weld_entries:
            ws.append([
                entry.iso_number,
                entry.weld_number,
                entry.evaluation,
                report.inspection_date.strftime('%m/%d/%Y') if report.inspection_date else '',
                report.job_number,
                report.project_name,
                report.client,
                report.location,
                report.specification,
                report.inspector_name,
                report.inspector_cert,
            ])

    for col in ws.columns:
        max_length = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)


def _create_ut_sheet(wb, filters):
    """Create UT-specific sheet with indication data."""
    ut_filters = dict(filters or {})
    ut_filters['inspection_type'] = 'UT'
    reports = _build_query(ut_filters).all()

    if not reports:
        return

    ws = wb.create_sheet('UT Reports')
    headers = [
        'ISO #', 'Weld #', 'Result', 'Date', 'Job Number', 'Project',
        'Material', 'Weld Process', 'Thickness', 'Joint Design',
        'Probe Angle', 'Ref Level (dB)', 'Inspector'
    ]
    ws.append(headers)
    _style_header(ws)

    for report in reports:
        ut_detail = report.ut_details
        for entry in report.weld_entries:
            ut_ind = entry.ut_indication
            ws.append([
                entry.iso_number,
                entry.weld_number,
                entry.evaluation,
                report.inspection_date.strftime('%m/%d/%Y') if report.inspection_date else '',
                report.job_number,
                report.project_name,
                ut_detail.material_type if ut_detail else '',
                ut_detail.weld_process if ut_detail else '',
                ut_detail.material_thickness if ut_detail else '',
                ut_detail.weld_joint_design if ut_detail else '',
                ut_ind.probe_angle if ut_ind else '',
                ut_ind.reference_level if ut_ind else '',
                report.inspector_name,
            ])

    for col in ws.columns:
        max_length = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)
