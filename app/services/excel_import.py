from openpyxl import load_workbook
from app import db
from app.models import InspectionReport, WeldEntry


def import_nde_dashboard(xlsx_path):
    """Import the existing NDE Dashboard Combined Excel file as seed data.

    Reads each sheet and attempts to map columns to the database schema.
    Returns a summary dict with counts.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    summary = {'sheets_processed': 0, 'reports_created': 0, 'entries_created': 0}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            continue

        headers = [str(h).strip().lower() if h else '' for h in rows[0]]
        col_map = _detect_columns(headers)

        if not col_map:
            continue

        summary['sheets_processed'] += 1

        # Create a generic report for this sheet's data
        report = InspectionReport(
            job_number=f'IMPORT-{sheet_name}',
            inspection_type='IMPORTED',
            project_name=sheet_name,
            result_summary='Imported',
            pdf_filename=f'Imported from {xlsx_path}',
        )
        db.session.add(report)
        db.session.flush()
        summary['reports_created'] += 1

        for row in rows[1:]:
            if not row or all(cell is None for cell in row):
                continue

            entry = _create_entry_from_row(row, col_map, report.id)
            if entry:
                db.session.add(entry)
                summary['entries_created'] += 1

    db.session.commit()
    wb.close()
    return summary


def _detect_columns(headers):
    """Auto-detect column mapping from header names."""
    col_map = {}
    iso_keywords = ['iso', 'iso #', 'iso#', 'indication', 'weld id', 'joint', 'joint #']
    weld_keywords = ['weld', 'weld #', 'weld no', 'weld number']
    result_keywords = ['result', 'status', 'evaluation', 'accept', 'vt', 'ut']

    for i, header in enumerate(headers):
        h = header.lower().strip()
        if any(k in h for k in iso_keywords):
            col_map.setdefault('iso', i)
        elif any(k in h for k in weld_keywords):
            col_map.setdefault('weld', i)
        elif any(k in h for k in result_keywords):
            col_map.setdefault('result', i)

    # Must have at least an ISO column to be useful
    return col_map if 'iso' in col_map else None


def _create_entry_from_row(row, col_map, report_id):
    """Create a WeldEntry from an Excel row using the column map."""
    iso = str(row[col_map['iso']]).strip() if col_map.get('iso') is not None and row[col_map['iso']] else None

    if not iso:
        return None

    weld_num = ''
    if col_map.get('weld') is not None and len(row) > col_map['weld'] and row[col_map['weld']]:
        weld_num = str(row[col_map['weld']]).strip()

    evaluation = ''
    if col_map.get('result') is not None and len(row) > col_map['result'] and row[col_map['result']]:
        evaluation = str(row[col_map['result']]).strip()

    return WeldEntry(
        report_id=report_id,
        iso_number=iso,
        weld_number=weld_num,
        evaluation=evaluation,
    )
